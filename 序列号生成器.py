# -*- coding: utf-8 -*-
# @Time    : 2024-5-24 13:43
# @Author  : Kali
# @FileName: 序列号生成器.py
# @Software: PyCharm
import os
import sys
from datetime import datetime

import pandas as pd
from PyQt5.QtCore import pyqtSlot
from openpyxl import Workbook
from openpyxl import load_workbook
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QMainWindow, QApplication, QMessageBox, QTableWidgetItem
from openpyxl.styles import PatternFill

from UI.untitled_vin import Ui_MainWindow
from Config.config import ConfigVin


# 主程序
class VehicleTest(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.init_ui()

    # 初始化界面
    def init_ui(self):
        # 标题
        self.setWindowTitle('序列号码生成器v1.0.0')
        self.setWindowIcon(QIcon('../image/logo.jpg'))
        self.data_initialize()
        self.control_ui()
        # 初始化标志
        self.reset_flags()
        # 连接信号
        self.connect_signals()

    def reset_flags(self):
        self.first_time_subtractVIN = False
        self.first_time_subtractMotor_serial_number = False
        self.first_time_subtractPCBA = False

    def connect_signals(self):
        self.lineEdit_5.textChanged.connect(lambda text: self.checkLineEditText(text, 'VIN'))
        self.lineEdit_2.textChanged.connect(lambda text: self.checkLineEditText(text, 'Motor_serial_number'))
        self.lineEdit_8.textChanged.connect(lambda text: self.checkLineEditText(text, 'PCBA'))

    @pyqtSlot(str)
    def checkLineEditText(self, text, flag_name):
        # 检查文本是否为空
        if text == '':
            # 如果文本为空，将标志重置为 False
            setattr(self, f'first_time_subtract{flag_name}', False)

    # 数据初始化
    def data_initialize(self):
        filename = "./excel/db.xlsx"
        # 判断文件是否存在
        if not os.path.exists(filename):
            QMessageBox.warning(self, '警告', '初始化文件不存在！')
            return
        # 加载工作簿
        workbook = load_workbook(filename=filename)
        # 选择活动工作表
        sheet = workbook.active
        # 读取特定单元格
        cell_value_A1 = sheet["A1"].value
        cell_value_B1 = sheet["B1"].value
        cell_value_C1 = sheet["C1"].value
        # 检查单元格值是否为空以及长度是否为17
        if not cell_value_A1 or len(cell_value_A1) != 17:
            QMessageBox.warning(self, '警告', 'VIN号初始化错误！')
            return
        if not cell_value_B1 or len(cell_value_B1) != 17:
            QMessageBox.warning(self, '警告', '电机序列号初始化错误！')
            return
        if not cell_value_C1 or len(cell_value_C1) != 17:
            QMessageBox.warning(self, '警告', 'PCBA序列号初始化错误！')
            return

        #  将VIN号，追加到输入框里面，更新输入框的值
        input1, input2, input3, input4, input5, input6, input7, input8, input9, input10, input11 \
            , input12, input13, input14, input15, input16, input17 = list(cell_value_A1)
        self.comboBox_9.setCurrentText(str(input1))
        self.comboBox_12.setCurrentText(str(input2))
        self.comboBox_13.setCurrentText(str(input3))
        self.comboBox_14.setCurrentText(str(input4))
        self.comboBox_15.setCurrentText(str(input5))
        self.comboBox_16.setCurrentText(str(input6))
        self.comboBox_17.setCurrentText(str(input7))
        self.comboBox_18.setCurrentText(str(input8))
        self.lineEdit_4.setText(str(input9))
        self.comboBox_19.setCurrentText(str(input10))
        self.comboBox_20.setCurrentText(str(input11))
        self.lineEdit_5.setText(str(input12 + input13 + input14 + input15 + input16 + input17))

        #  将电机序列号，追加到输入框里面，更新输入框的值
        input1, input2, input3, input4, input5, input6, input7, input8, input9, input10, input11 \
            , input12, input13, input14, input15, input16, input17 = list(cell_value_B1)
        self.comboBox.setCurrentText(str(input1))
        self.comboBox_3.setCurrentText(str(input2))
        self.comboBox_2.setCurrentText(str(input3))
        self.comboBox_4.setCurrentText(str(input4))
        self.comboBox_5.setCurrentText(str(input5))
        self.comboBox_6.setCurrentText(str(input6))
        self.comboBox_7.setCurrentText(str(input7))
        self.comboBox_8.setCurrentText(str(input8))
        self.lineEdit.setText(str(input9))
        self.comboBox_10.setCurrentText(str(input10))
        self.comboBox_11.setCurrentText(str(input11))
        self.lineEdit_2.setText(str(input12 + input13 + input14 + input15 + input16 + input17))

        #  将PCB A序列号，追加到输入框里面，更新输入框的值
        input1, input2, input3, input4, input5, input6, input7, input8, input9, input10, input11 \
            , input12, input13, input14, input15, input16, input17 = list(cell_value_C1)
        self.comboBox_21.setCurrentText(str(input1))
        self.comboBox_22.setCurrentText(str(input2))
        self.comboBox_23.setCurrentText(str(input3))
        self.comboBox_24.setCurrentText(str(input4))
        self.comboBox_25.setCurrentText(str(input5))
        self.comboBox_26.setCurrentText(str(input6))
        self.comboBox_27.setCurrentText(str(input7))
        self.comboBox_28.setCurrentText(str(input8))
        self.lineEdit_7.setText(str(input9))
        self.comboBox_29.setCurrentText(str(input10))
        self.comboBox_30.setCurrentText(str(input11))
        self.lineEdit_8.setText(str(input12 + input13 + input14 + input15 + input16 + input17))

    def control_ui(self):
        #  手动生成VIN
        self.pushButton_5.clicked.connect(self.ManuallyGenerateVIN)
        #  自动生成VIN
        self.pushButton_6.clicked.connect(self.AutomaticallyGenerateVIN)
        # # 导出Excel
        self.pushButton_7.clicked.connect(self.ExportExcel_VIN)
        # # 清除表格
        self.pushButton_8.clicked.connect(self.clear_tableWidget_2)

        #  手动生成电机序列号
        self.pushButton.clicked.connect(self.Motor_serial_number)
        #  自动生成电机序列号
        self.pushButton_2.clicked.connect(self.AutomaticallyGenerateMotor_serial_number)
        # # 导出Excel 电机序列号
        self.pushButton_4.clicked.connect(self.ExportExcel_Motor_serial_number)
        # # 清除表格
        self.pushButton_3.clicked.connect(self.clear_tableWidget)

        #  手动生成PCBA
        self.pushButton_10.clicked.connect(self.Manually_generate_PCBA)
        #  自动生成PCBA
        self.pushButton_11.clicked.connect(self.The_PCBA_is_automatically_generated)
        # # 导出Excel PCBA
        self.pushButton_12.clicked.connect(self.ExportExcel_PCBA)
        # # 清除表格
        self.pushButton_9.clicked.connect(self.clear_tableWidget_3)

    #  手动生成VIN
    def ManuallyGenerateVIN(self):
        V1 = self.comboBox_9.currentText()  # V1
        V2 = self.comboBox_12.currentText()  # V2
        V3 = self.comboBox_13.currentText()  # V3
        V4 = self.comboBox_14.currentText()  # V4
        V5 = self.comboBox_15.currentText()  # V5
        V6 = self.comboBox_16.currentText()  # V6
        V7 = self.comboBox_17.currentText()  # V7
        V8 = self.comboBox_18.currentText()  # V8
        # V9 = self.lineEdit_7.text()  # V9
        V10 = self.comboBox_19.currentText()  # V10
        V11 = self.comboBox_20.currentText()  # V11
        V12 = self.lineEdit_5.text()  # V12~V17

        # 检查所有字段是否都不为空
        if '' in [V1, V2, V3, V4, V5, V6, V7, V8, V10, V11, V12, ]:
            QMessageBox.warning(self, "错误", "所有字段都必须填写")
            return
        # 检查长度是否为6，不足则补0
        if len(V12) < 6:
            QMessageBox.warning(self, "错误", "V12不足为6位，请重新输入！")
            return
        # 检查长度是否为6，不足则补0
        if len(V12) > 6:
            QMessageBox.warning(self, "错误", "V12只能为6位，请重新输入！")
            return
        # 将其转换为整数
        V12toV17 = int(V12)
        # C检查最后六位数字是否已达到最大值
        if V12toV17 >= 999999:
            # 在没有翻译时执行
            QMessageBox.warning(self, "错误", "VIN的最后六位已达到最大值999999，无法再增加")
            return

        # 将其转换回字符串，必要时用零填充
        last_six_str = str(V12toV17).zfill(6)
        V1_C = int(ConfigVin.content_map.get(V1))
        V2_C = int(ConfigVin.content_map.get(V2))
        V3_C = int(ConfigVin.content_map.get(V3))
        V4_C = int(ConfigVin.content_map.get(V4))
        V5_C = int(ConfigVin.content_map.get(V5))
        V6_C = int(ConfigVin.content_map.get(V6))
        V7_C = int(ConfigVin.content_map.get(V7))
        V8_C = int(ConfigVin.content_map.get(V8))
        V10_C = int(ConfigVin.content_map.get(V10))
        V11_C = int(ConfigVin.content_map.get(V11))
        # 计算出VIN号
        VinN = (V1_C * 8 + V2_C * 7 + V3_C * 6 + V4_C * 5 + V5_C * 4 + V6_C * 3 +
                V7_C * 2 + V8_C * 10 + V10_C * 9 + V11_C * 8 + int(last_six_str[0]) * 7 + int(last_six_str[1]) * 6 + int(last_six_str[2]) * 5 +
                int(last_six_str[3]) * 4 + int(last_six_str[4]) * 3 + int(last_six_str[5]) * 2)
        vin9 = VinN % 11
        if vin9 == 10:
            vin9 = "X"
        self.lineEdit_4.setText(str(vin9))
        VIN = V1 + V2 + V3 + V4 + V5 + V6 + V7 + V8 + str(vin9) + V10 + V11 + V12

        # 将VIN的后六位数字替换为增加的值
        VIN = VIN[:-6] + last_six_str
        # 插入到self.tableWidget_2
        rowPosition = self.tableWidget_2.rowCount()
        self.tableWidget_2.insertRow(rowPosition)
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        id_item = QTableWidgetItem(str(rowPosition + 1))
        id_item.setTextAlignment(0x0004 | 0x0080)
        self.tableWidget_2.setItem(rowPosition, 0, id_item)  # ID

        vin_item = QTableWidgetItem(VIN)
        vin_item.setTextAlignment(0x0004 | 0x0080)
        self.tableWidget_2.setItem(rowPosition, 1, vin_item)  # VIN

        time_item = QTableWidgetItem(current_time)
        time_item.setTextAlignment(0x0004 | 0x0080)
        self.tableWidget_2.setItem(rowPosition, 2, time_item)  # 创建时间

        last_six_digits = VIN[-6:]
        last_six_digits_int = int(last_six_digits)
        # 增加整数
        last_six_digits_int += 1
        # 将其转换回字符串，必要时用零填充
        last_six_str = str(last_six_digits_int).zfill(6)
        #  将VIN号，追加到输入框里面，更新输入框的值
        self.lineEdit_5.setText(last_six_str)
        # 检查文件是否存在
        filename = "./excel/db.xlsx"
        try:
            if os.path.exists(filename):
                # 文件存在，加载工作簿
                workbook = load_workbook(filename=filename)
            else:
                # 文件不存在，创建一个新的工作簿
                workbook = Workbook()

            # 选择要写入的工作表
            sheet = workbook.active
            # 写入特定单元格
            sheet["A1"] = VIN
            # 保存工作簿
            workbook.save(filename=filename)
        except PermissionError:
            # 提示用户文件无法写入
            QMessageBox.warning(self, "错误", "文件已经被打开，无法写入。请关闭文件后重试。")

    #  自动生成VIN
    def AutomaticallyGenerateVIN(self):
        # 获取循环的次数
        lineEdit_6 = self.lineEdit_6.text()
        # 设置只能输入数字，且只能输入小于2000的数字，需要判断numb 值是否小于等于2000
        if not lineEdit_6.isdigit() or int(lineEdit_6) > 2000:
            # 在没有翻译时执行
            QMessageBox.warning(self, "错误", "无效的输入。请输入一个小于或等于2000的数字。")
            return

        numb = int(lineEdit_6)
        V1 = self.comboBox_9.currentText()  # V1
        V2 = self.comboBox_12.currentText()  # V2
        V3 = self.comboBox_13.currentText()  # V3
        V4 = self.comboBox_14.currentText()  # V4
        V5 = self.comboBox_15.currentText()  # V5
        V6 = self.comboBox_16.currentText()  # V6
        V7 = self.comboBox_17.currentText()  # V7
        V8 = self.comboBox_18.currentText()  # V8
        # V9 = self.lineEdit_7.text()  # V9
        V10 = self.comboBox_19.currentText()  # V10
        V11 = self.comboBox_20.currentText()  # V11
        # 获取V12 - V17输入框的值
        V12 = self.lineEdit_5.text()  # V12~V17

        # 检查所有字段是否都不为空
        if '' in [V1, V2, V3, V4, V5, V6, V7, V8, V10, V11, V12]:
            # 在没有翻译时执行
            QMessageBox.warning(self, "错误", "所有字段都必须填写")
            return
        # 检查长度是否为6，不足则补0
        if len(V12) < 6:
            QMessageBox.warning(self, "错误", "V12不足为6位，请重新输入！")
            return
        # 检查长度是否为6，不足则补0
        if len(V12) > 6:
            QMessageBox.warning(self, "错误", "V12只能为6位，请重新输入！")
            return
        # 将其转换为整数
        last_six_int = int(V12)
        # 检查是否已经进行过减一操作
        if not self.first_time_subtractVIN:
            last_six_int -= 1
            self.first_time_subtractVIN = True
        # 初始化VIN_data
        VIN_data = ''
        VIN = str(last_six_int).zfill(6)  # 将V12转换为字符串并在前面填充0，直到长度为6

        # 循环numb 次数
        for _ in range(numb):
            VIN = str(int(VIN) + 1).zfill(6)
            V1_C = int(ConfigVin.content_map.get(V1))
            V2_C = int(ConfigVin.content_map.get(V2))
            V3_C = int(ConfigVin.content_map.get(V3))
            V4_C = int(ConfigVin.content_map.get(V4))
            V5_C = int(ConfigVin.content_map.get(V5))
            V6_C = int(ConfigVin.content_map.get(V6))
            V7_C = int(ConfigVin.content_map.get(V7))
            V8_C = int(ConfigVin.content_map.get(V8))
            V10_C = int(ConfigVin.content_map.get(V10))
            V11_C = int(ConfigVin.content_map.get(V11))
            # 计算出VIN号
            VinN = (V1_C * 8 + V2_C * 7 + V3_C * 6 + V4_C * 5 + V5_C * 4 + V6_C * 3 +
                    V7_C * 2 + V8_C * 10 + V10_C * 9 + V11_C * 8 + int(VIN[0]) * 7 + int(VIN[1]) * 6 + int(VIN[2]) * 5 +
                    int(VIN[3]) * 4 + int(VIN[4]) * 3 + int(VIN[5]) * 2)
            vin9 = VinN % 11
            if vin9 == 10:
                vin9 = "X"
            #  将VIN号，追加到输入框里面，更新输入框的值
            self.lineEdit_4.setText(str(vin9))
            self.lineEdit_5.setText(str(VIN))
            VIN_data = V1 + V2 + V3 + V4 + V5 + V6 + V7 + V8 + str(vin9) + V10 + V11 + VIN
            # 插入到self.tableWidget_2
            rowPosition = self.tableWidget_2.rowCount()
            self.tableWidget_2.insertRow(rowPosition)
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            id_item = QTableWidgetItem(str(rowPosition + 1))
            id_item.setTextAlignment(0x0004 | 0x0080)
            self.tableWidget_2.setItem(rowPosition, 0, id_item)  # ID

            vin_item = QTableWidgetItem(str(VIN_data))  # 将VIN转换为字符串
            vin_item.setTextAlignment(0x0004 | 0x0080)
            self.tableWidget_2.setItem(rowPosition, 1, vin_item)  # VIN

            time_item = QTableWidgetItem(current_time)
            time_item.setTextAlignment(0x0004 | 0x0080)
            self.tableWidget_2.setItem(rowPosition, 2, time_item)  # 创建时间
            # 在没有翻译时执行
        # 检查文件是否存在
        filename = "./excel/db.xlsx"
        try:
            if os.path.exists(filename):
                # 文件存在，加载工作簿
                workbook = load_workbook(filename=filename)
            else:
                # 文件不存在，创建一个新的工作簿
                workbook = Workbook()

            # 选择要写入的工作表
            sheet = workbook.active
            # 写入特定单元格
            sheet["A1"] = VIN_data
            # 保存工作簿
            workbook.save(filename=filename)
        except PermissionError:
            # 提示用户文件无法写入
            QMessageBox.warning(self, "错误", "文件已经被打开，无法写入。请关闭文件后重试。")

    # 导出Excel VIN
    def ExportExcel_VIN(self):
        # 获取表中的行数和列数
        rows = self.tableWidget_2.rowCount()
        cols = self.tableWidget_2.columnCount()

        # 检查表格是否为空
        if rows == 0 or cols == 0:
            QMessageBox.about(self, "提示", "表格为空，无法导出！")
            return

        # 表头
        headers = ['ID', 'VIN', '创建时间']

        # 创建一个DataFrame来存储表数据
        data = []
        for row in range(rows):
            row_data = []
            for col in range(cols):
                item = self.tableWidget_2.item(row, col)
                if item is not None:
                    row_data.append(item.text())
                else:
                    row_data.append('')
            data.append(row_data)

        # 将数据转换为DataFrame
        df = pd.DataFrame(data, columns=headers)

        # 获取当前时间并格式化为字符串
        current_time = datetime.now().strftime("%Y%m%d%H%M%S")

        # 通过添加当前时间来创建文件名
        file_name = f"VIN_{current_time}.xlsx"

        # 获取桌面路径
        desktop_path = os.path.expanduser("~/Desktop")

        # 创建文件的完整路径
        file_path = os.path.join(desktop_path, file_name)

        # 将DataFrame写入Excel文件
        df.to_excel(file_path, index=False)

        # 使用openpyxl加载生成的Excel文件
        wb = load_workbook(file_path)
        ws = wb.active

        # 固定第一行
        ws.freeze_panes = ws['A2']

        # 设置表头颜色为浅绿色
        header_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # 保存修改后的文件
        wb.save(file_path)

        # 弹出消息框提示用户文件已保存
        QMessageBox.about(self, "提示", f"文件已保存到桌面：{file_name}")

    #  清除tableWidget_2
    @pyqtSlot()
    def clear_tableWidget_2(self):
        self.tableWidget_2.setRowCount(0)

    #  手动生成电机序列号
    def Motor_serial_number(self):
        V1 = self.comboBox.currentText()  # V1
        V2 = self.comboBox_3.currentText()  # V2
        V3 = self.comboBox_2.currentText()  # V3
        V4 = self.comboBox_4.currentText()  # V4
        V5 = self.comboBox_5.currentText()  # V5
        V6 = self.comboBox_6.currentText()  # V6
        V7 = self.comboBox_7.currentText()  # V7
        V8 = self.comboBox_8.currentText()  # V8
        V10 = self.comboBox_10.currentText()  # V10
        V11 = self.comboBox_11.currentText()  # V11
        V12 = self.lineEdit_2.text()  # V12~V17

        # 检查所有字段是否都不为空
        if '' in [V1, V2, V3, V4, V5, V6, V7, V8, V10, V11, V12, ]:
            QMessageBox.warning(self, "错误", "所有字段都必须填写")
            return
        # 检查长度是否为6，不足则补0
        if len(V12) < 6:
            QMessageBox.warning(self, "错误", "V12不足为6位，请重新输入！")
            return
        # 检查长度是否为6，不足则补0
        if len(V12) > 6:
            QMessageBox.warning(self, "错误", "V12只能为6位，请重新输入！")
            return
        # 将其转换为整数
        V12toV17 = int(V12)
        # C检查最后六位数字是否已达到最大值
        if V12toV17 >= 999999:
            # 在没有翻译时执行
            QMessageBox.warning(self, "错误", "VIN的最后六位已达到最大值999999，无法再增加")
            return

        # 将其转换回字符串，必要时用零填充
        last_six_str = str(V12toV17).zfill(6)
        V1_C = int(ConfigVin.content_map.get(V1))
        V2_C = int(ConfigVin.content_map.get(V2))
        V3_C = int(ConfigVin.content_map.get(V3))
        V4_C = int(ConfigVin.content_map.get(V4))
        V5_C = int(ConfigVin.content_map.get(V5))
        V6_C = int(ConfigVin.content_map.get(V6))
        V7_C = int(ConfigVin.content_map.get(V7))
        V8_C = int(ConfigVin.content_map.get(V8))
        V10_C = int(ConfigVin.content_map.get(V10))
        V11_C = int(ConfigVin.content_map.get(V11))
        # 计算出VIN号
        VinN = (V1_C * 8 + V2_C * 7 + V3_C * 6 + V4_C * 5 + V5_C * 4 + V6_C * 3 +
                V7_C * 2 + V8_C * 10 + V10_C * 9 + V11_C * 8 + int(V12[0]) * 7 + int(V12[1]) * 6 + int(V12[2]) * 5 +
                int(V12[3]) * 4 + int(V12[4]) * 3 + int(V12[5]) * 2)
        vin9 = VinN % 11
        if vin9 == 10:
            vin9 = "X"
        self.lineEdit.setText(str(vin9))
        VIN = V1 + V2 + V3 + V4 + V5 + V6 + V7 + V8 + str(vin9) + V10 + V11 + V12
        # 将VIN的后六位数字替换为增加的值
        VIN = VIN[:-6] + last_six_str
        # 插入到self.tableWidget
        rowPosition = self.tableWidget.rowCount()
        self.tableWidget.insertRow(rowPosition)
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        id_item = QTableWidgetItem(str(rowPosition + 1))
        id_item.setTextAlignment(0x0004 | 0x0080)
        self.tableWidget.setItem(rowPosition, 0, id_item)  # ID

        vin_item = QTableWidgetItem(VIN)
        vin_item.setTextAlignment(0x0004 | 0x0080)
        self.tableWidget.setItem(rowPosition, 1, vin_item)  # VIN

        time_item = QTableWidgetItem(current_time)
        time_item.setTextAlignment(0x0004 | 0x0080)
        self.tableWidget.setItem(rowPosition, 2, time_item)  # 创建时间

        last_six_digits = VIN[-6:]
        last_six_digits_int = int(last_six_digits)
        # 增加整数
        last_six_digits_int += 1
        # 将其转换回字符串，必要时用零填充
        last_six_str = str(last_six_digits_int).zfill(6)
        #  将VIN号，追加到输入框里面，更新输入框的值
        self.lineEdit_2.setText(last_six_str)
        # 检查文件是否存在
        filename = "./excel/db.xlsx"
        try:
            if os.path.exists(filename):
                # 文件存在，加载工作簿
                workbook = load_workbook(filename=filename)
            else:
                # 文件不存在，创建一个新的工作簿
                workbook = Workbook()

            # 选择要写入的工作表
            sheet = workbook.active
            # 写入特定单元格
            sheet["B1"] = VIN
            # 保存工作簿
            workbook.save(filename=filename)
        except PermissionError:
            # 提示用户文件无法写入
            QMessageBox.warning(self, "错误", "文件已经被打开，无法写入。请关闭文件后重试。")

    #  自动生成电机序列号
    def AutomaticallyGenerateMotor_serial_number(self):
        # 获取循环的次数
        lineEdit_3 = self.lineEdit_3.text()
        # 设置只能输入数字，且只能输入小于2000的数字，需要判断numb 值是否小于等于2000
        if not lineEdit_3.isdigit() or int(lineEdit_3) > 2000:
            # 在没有翻译时执行
            QMessageBox.warning(self, "错误", "无效的输入。请输入一个小于或等于2000的数字。")
            return

        numb = int(lineEdit_3)
        V1 = self.comboBox.currentText()  # V1
        V2 = self.comboBox_3.currentText()  # V2
        V3 = self.comboBox_2.currentText()  # V3
        V4 = self.comboBox_4.currentText()  # V4
        V5 = self.comboBox_5.currentText()  # V5
        V6 = self.comboBox_6.currentText()  # V6
        V7 = self.comboBox_7.currentText()  # V7
        V8 = self.comboBox_8.currentText()  # V8
        V10 = self.comboBox_10.currentText()  # V10
        V11 = self.comboBox_11.currentText()  # V11
        V12 = self.lineEdit_2.text()  # V12~V17

        # 检查所有字段是否都不为空
        if '' in [V1, V2, V3, V4, V5, V6, V7, V8, V10, V11, V12]:
            # 在没有翻译时执行
            QMessageBox.warning(self, "错误", "所有字段都必须填写")
            return
        # 检查长度是否为6，不足则补0
        if len(V12) < 6:
            QMessageBox.warning(self, "错误", "V12不足为6位，请重新输入！")
            return
        # 检查长度是否为6，不足则补0
        if len(V12) > 6:
            QMessageBox.warning(self, "错误", "V12只能为6位，请重新输入！")
            return
        # 将其转换为整数
        last_six_int = int(V12)
        # 检查是否已经进行过减一操作
        if not self.first_time_subtractMotor_serial_number:
            last_six_int -= 1
            self.first_time_subtractMotor_serial_number = True
        # 初始化VIN_data
        VIN_data = ''
        VIN = str(last_six_int).zfill(6)  # 将V12转换为字符串并在前面填充0，直到长度为6

        # 循环numb 次数
        for _ in range(numb):
            VIN = str(int(VIN) + 1).zfill(6)
            V1_C = int(ConfigVin.content_map.get(V1))
            V2_C = int(ConfigVin.content_map.get(V2))
            V3_C = int(ConfigVin.content_map.get(V3))
            V4_C = int(ConfigVin.content_map.get(V4))
            V5_C = int(ConfigVin.content_map.get(V5))
            V6_C = int(ConfigVin.content_map.get(V6))
            V7_C = int(ConfigVin.content_map.get(V7))
            V8_C = int(ConfigVin.content_map.get(V8))
            V10_C = int(ConfigVin.content_map.get(V10))
            V11_C = int(ConfigVin.content_map.get(V11))
            # 计算出VIN号
            VinN = (V1_C * 8 + V2_C * 7 + V3_C * 6 + V4_C * 5 + V5_C * 4 + V6_C * 3 +
                    V7_C * 2 + V8_C * 10 + V10_C * 9 + V11_C * 8 + int(VIN[0]) * 7 + int(VIN[1]) * 6 + int(VIN[2]) * 5 +
                    int(VIN[3]) * 4 + int(VIN[4]) * 3 + int(VIN[5]) * 2)
            vin9 = VinN % 11
            if vin9 == 10:
                vin9 = "X"
            #  将VIN号，追加到输入框里面，更新输入框的值
            self.lineEdit.setText(str(vin9))
            self.lineEdit_2.setText(str(VIN))
            VIN_data = V1 + V2 + V3 + V4 + V5 + V6 + V7 + V8 + str(vin9) + V10 + V11 + VIN
            # 插入到self.tableWidget
            rowPosition = self.tableWidget.rowCount()
            self.tableWidget.insertRow(rowPosition)
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            id_item = QTableWidgetItem(str(rowPosition + 1))
            id_item.setTextAlignment(0x0004 | 0x0080)
            self.tableWidget.setItem(rowPosition, 0, id_item)  # ID

            vin_item = QTableWidgetItem(str(VIN_data))  # 将VIN转换为字符串
            vin_item.setTextAlignment(0x0004 | 0x0080)
            self.tableWidget.setItem(rowPosition, 1, vin_item)  # VIN

            time_item = QTableWidgetItem(current_time)
            time_item.setTextAlignment(0x0004 | 0x0080)
            self.tableWidget.setItem(rowPosition, 2, time_item)  # 创建时间
            # 在没有翻译时执行
        # 检查文件是否存在
        filename = "./excel/db.xlsx"
        try:
            if os.path.exists(filename):
                # 文件存在，加载工作簿
                workbook = load_workbook(filename=filename)
            else:
                # 文件不存在，创建一个新的工作簿
                workbook = Workbook()

            # 选择要写入的工作表
            sheet = workbook.active
            # 写入特定单元格
            sheet["B1"] = VIN_data
            # 保存工作簿
            workbook.save(filename=filename)
        except PermissionError:
            # 提示用户文件无法写入
            QMessageBox.warning(self, "错误", "文件已经被打开，无法写入。请关闭文件后重试。")

    # 导出Excel  电机序列号
    def ExportExcel_Motor_serial_number(self):
        # 获取表中的行数和列数
        rows = self.tableWidget.rowCount()
        cols = self.tableWidget.columnCount()

        # 检查表格是否为空
        if rows == 0 or cols == 0:
            QMessageBox.about(self, "提示", "表格为空，无法导出！")
            return

        # 表头
        headers = ['ID', '电机序列号', '创建时间']

        # 创建一个DataFrame来存储表数据
        data = []
        for row in range(rows):
            row_data = []
            for col in range(cols):
                item = self.tableWidget.item(row, col)
                if item is not None:
                    row_data.append(item.text())
                else:
                    row_data.append('')
            data.append(row_data)

        # 将数据转换为DataFrame
        df = pd.DataFrame(data, columns=headers)

        # 获取当前时间并格式化为字符串
        current_time = datetime.now().strftime("%Y%m%d%H%M%S")

        # 通过添加当前时间来创建文件名
        file_name = f"Motor_serial_number-{current_time}.xlsx"

        # 获取桌面路径
        desktop_path = os.path.expanduser("~/Desktop")

        # 创建文件的完整路径
        file_path = os.path.join(desktop_path, file_name)

        # 将DataFrame写入Excel文件
        df.to_excel(file_path, index=False)

        # 使用openpyxl加载生成的Excel文件
        wb = load_workbook(file_path)
        ws = wb.active

        # 固定第一行
        ws.freeze_panes = ws['A2']

        # 设置表头颜色为浅绿色
        header_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # 保存修改后的文件
        wb.save(file_path)

        # 弹出消息框提示用户文件已保存
        QMessageBox.about(self, "提示", f"文件已保存到桌面：{file_name}")

    #  清除tableWidget
    @pyqtSlot()
    def clear_tableWidget(self):
        self.tableWidget.setRowCount(0)

    #  手动生成PCBA
    def Manually_generate_PCBA(self):
        V1 = self.comboBox_21.currentText()  # V1
        V2 = self.comboBox_22.currentText()  # V2
        V3 = self.comboBox_23.currentText()  # V3
        V4 = self.comboBox_24.currentText()  # V4
        V5 = self.comboBox_25.currentText()  # V5
        V6 = self.comboBox_26.currentText()  # V6
        V7 = self.comboBox_27.currentText()  # V7
        V8 = self.comboBox_28.currentText()  # V8
        V10 = self.comboBox_29.currentText()  # V10
        V11 = self.comboBox_30.currentText()  # V11
        V12 = self.lineEdit_8.text()  # V12~V17

        # 检查所有字段是否都不为空
        if '' in [V1, V2, V3, V4, V5, V6, V7, V8, V10, V11, V12, ]:
            QMessageBox.warning(self, "错误", "所有字段都必须填写")
            return
        # 检查长度是否为6，不足则补0
        if len(V12) < 6:
            QMessageBox.warning(self, "错误", "V12不足为6位，请重新输入！")
            return
        # 检查长度是否为6，不足则补0
        if len(V12) > 6:
            QMessageBox.warning(self, "错误", "V12只能为6位，请重新输入！")
            return
        # 将其转换为整数
        V12toV17 = int(V12)
        # C检查最后六位数字是否已达到最大值
        if V12toV17 >= 999999:
            # 在没有翻译时执行
            QMessageBox.warning(self, "错误", "VIN的最后六位已达到最大值999999，无法再增加")
            return

        # 将其转换回字符串，必要时用零填充
        last_six_str = str(V12toV17).zfill(6)

        V1_C = int(ConfigVin.content_map.get(V1))
        V2_C = int(ConfigVin.content_map.get(V2))
        V3_C = int(ConfigVin.content_map.get(V3))
        V4_C = int(ConfigVin.content_map.get(V4))
        V5_C = int(ConfigVin.content_map.get(V5))
        V6_C = int(ConfigVin.content_map.get(V6))
        V7_C = int(ConfigVin.content_map.get(V7))
        V8_C = int(ConfigVin.content_map.get(V8))
        V10_C = int(ConfigVin.content_map.get(V10))
        V11_C = int(ConfigVin.content_map.get(V11))
        # 计算出VIN号
        VinN = (V1_C * 8 + V2_C * 7 + V3_C * 6 + V4_C * 5 + V5_C * 4 + V6_C * 3 +
                V7_C * 2 + V8_C * 10 + V10_C * 9 + V11_C * 8 + int(last_six_str[0]) * 7 + int(last_six_str[1]) * 6 + int(last_six_str[2]) * 5 +
                int(last_six_str[3]) * 4 + int(last_six_str[4]) * 3 + int(last_six_str[5]) * 2)
        vin9 = VinN % 11
        if vin9 == 10:
            vin9 = "X"
        self.lineEdit_7.setText(str(vin9))
        VIN = V1 + V2 + V3 + V4 + V5 + V6 + V7 + V8 + str(vin9) + V10 + V11 + V12
        # 将VIN的后六位数字替换为增加的值
        VIN = VIN[:-6] + last_six_str
        # 插入到self.tableWidget_3
        rowPosition = self.tableWidget_3.rowCount()
        self.tableWidget_3.insertRow(rowPosition)
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        id_item = QTableWidgetItem(str(rowPosition + 1))
        id_item.setTextAlignment(0x0004 | 0x0080)
        self.tableWidget_3.setItem(rowPosition, 0, id_item)  # ID

        vin_item = QTableWidgetItem(VIN)
        vin_item.setTextAlignment(0x0004 | 0x0080)
        self.tableWidget_3.setItem(rowPosition, 1, vin_item)  # VIN

        time_item = QTableWidgetItem(current_time)
        time_item.setTextAlignment(0x0004 | 0x0080)
        self.tableWidget_3.setItem(rowPosition, 2, time_item)  # 创建时间

        last_six_digits = VIN[-6:]
        last_six_digits_int = int(last_six_digits)
        # 增加整数
        last_six_digits_int += 1
        # 将其转换回字符串，必要时用零填充
        last_six_str = str(last_six_digits_int).zfill(6)
        #  将VIN号，追加到输入框里面，更新输入框的值
        self.lineEdit_8.setText(last_six_str)
        # 检查文件是否存在
        filename = "./excel/db.xlsx"
        try:
            if os.path.exists(filename):
                # 文件存在，加载工作簿
                workbook = load_workbook(filename=filename)
            else:
                # 文件不存在，创建一个新的工作簿
                workbook = Workbook()

            # 选择要写入的工作表
            sheet = workbook.active
            # 写入特定单元格
            sheet["C1"] = VIN
            # 保存工作簿
            workbook.save(filename=filename)
        except PermissionError:
            # 提示用户文件无法写入
            QMessageBox.warning(self, "错误", "文件已经被打开，无法写入。请关闭文件后重试。")

    #  自动生成PCBA
    def The_PCBA_is_automatically_generated(self):
        # 获取循环的次数
        numb = self.lineEdit_9.text()
        # 设置只能输入数字，且只能输入小于2000的数字，需要判断numb 值是否小于等于2000
        if not numb.isdigit() or int(numb) > 2000:
            # 在没有翻译时执行
            QMessageBox.warning(self, "错误", "无效的输入。请输入一个小于或等于2000的数字。")
            return

        numb = int(numb)
        V1 = self.comboBox_21.currentText()  # V1
        V2 = self.comboBox_22.currentText()  # V2
        V3 = self.comboBox_23.currentText()  # V3
        V4 = self.comboBox_24.currentText()  # V4
        V5 = self.comboBox_25.currentText()  # V5
        V6 = self.comboBox_26.currentText()  # V6
        V7 = self.comboBox_27.currentText()  # V7
        V8 = self.comboBox_28.currentText()  # V8
        # V9 = self.lineEdit_7.text()  # V9
        V10 = self.comboBox_29.currentText()  # V10
        V11 = self.comboBox_30.currentText()  # V11
        V12 = self.lineEdit_8.text()  # V12~V17

        # 检查所有字段是否都不为空
        if '' in [V1, V2, V3, V4, V5, V6, V7, V8, V10, V11, V12]:
            # 在没有翻译时执行
            QMessageBox.warning(self, "错误", "所有字段都必须填写")
            return
        # 检查长度是否为6，不足则补0
        if len(V12) < 6:
            QMessageBox.warning(self, "错误", "V12不足为6位，请重新输入！")
            return
        # 检查长度大于6位
        if len(V12) > 6:
            QMessageBox.warning(self, "错误", "V12只能为6位，请重新输入！")
            return
        # 将其转换为整数
        last_six_int = int(V12)
        # 检查是否已经进行过减一操作
        if not self.first_time_subtractPCBA:
            last_six_int -= 1
            self.first_time_subtractPCBA = True
        # 初始化VIN_data
        VIN_data = ''
        VIN = str(last_six_int).zfill(6)  # 将V12转换为字符串并在前面填充0，直到长度为6
        # 循环numb 次数
        for _ in range(numb):
            VIN = str(int(VIN) + 1).zfill(6)
            V1_C = int(ConfigVin.content_map.get(V1))
            V2_C = int(ConfigVin.content_map.get(V2))
            V3_C = int(ConfigVin.content_map.get(V3))
            V4_C = int(ConfigVin.content_map.get(V4))
            V5_C = int(ConfigVin.content_map.get(V5))
            V6_C = int(ConfigVin.content_map.get(V6))
            V7_C = int(ConfigVin.content_map.get(V7))
            V8_C = int(ConfigVin.content_map.get(V8))
            V10_C = int(ConfigVin.content_map.get(V10))
            V11_C = int(ConfigVin.content_map.get(V11))
            # 计算出VIN号
            VinN = (V1_C * 8 + V2_C * 7 + V3_C * 6 + V4_C * 5 + V5_C * 4 + V6_C * 3 +
                    V7_C * 2 + V8_C * 10 + V10_C * 9 + V11_C * 8 + int(VIN[0]) * 7 + int(VIN[1]) * 6 + int(
                        VIN[2]) * 5 +
                    int(VIN[3]) * 4 + int(VIN[4]) * 3 + int(VIN[5]) * 2)
            vin9 = VinN % 11
            if vin9 == 10:
                vin9 = "X"
            #  将PCBA号，追加到输入框里面，更新输入框的值
            self.lineEdit_7.setText(str(vin9))
            self.lineEdit_8.setText(str(VIN))
            VIN_data = V1 + V2 + V3 + V4 + V5 + V6 + V7 + V8 + str(vin9) + V10 + V11 + VIN
            # 插入到self.tableWidget_3
            rowPosition = self.tableWidget_3.rowCount()
            self.tableWidget_3.insertRow(rowPosition)
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            id_item = QTableWidgetItem(str(rowPosition + 1))
            id_item.setTextAlignment(0x0004 | 0x0080)
            self.tableWidget_3.setItem(rowPosition, 0, id_item)  # ID

            vin_item = QTableWidgetItem(str(VIN_data))  # 将VIN转换为字符串
            vin_item.setTextAlignment(0x0004 | 0x0080)
            self.tableWidget_3.setItem(rowPosition, 1, vin_item)  # VIN

            time_item = QTableWidgetItem(current_time)
            time_item.setTextAlignment(0x0004 | 0x0080)
            self.tableWidget_3.setItem(rowPosition, 2, time_item)  # 创建时间
        # 检查文件是否存在
        filename = "./excel/db.xlsx"
        try:
            if os.path.exists(filename):
                # 文件存在，加载工作簿
                workbook = load_workbook(filename=filename)
            else:
                # 文件不存在，创建一个新的工作簿
                workbook = Workbook()

            # 选择要写入的工作表
            sheet = workbook.active
            # 写入特定单元格
            sheet["C1"] = VIN_data
            # 保存工作簿
            workbook.save(filename=filename)
        except PermissionError:
            # 提示用户文件无法写入
            QMessageBox.warning(self, "错误", "文件已经被打开，无法写入。请关闭文件后重试。")

    # 导出Excel PCBA
    def ExportExcel_PCBA(self):

        # 获取表中的行数和列数
        rows = self.tableWidget_3.rowCount()
        cols = self.tableWidget_3.columnCount()

        # 检查表格是否为空
        if rows == 0 or cols == 0:
            QMessageBox.about(self, "提示", "表格为空，无法导出！")
            return

        # 表头
        headers = ['ID', 'PCBA序列号', '创建时间']

        # 创建一个DataFrame来存储表数据
        data = []
        for row in range(rows):
            row_data = []
            for col in range(cols):
                item = self.tableWidget_3.item(row, col)
                if item is not None:
                    row_data.append(item.text())
                else:
                    row_data.append('')
            data.append(row_data)

        # 将数据转换为DataFrame
        df = pd.DataFrame(data, columns=headers)

        # 获取当前时间并格式化为字符串
        current_time = datetime.now().strftime("%Y%m%d%H%M%S")

        # 通过添加当前时间来创建文件名
        file_name = f"PCBA_{current_time}.xlsx"

        # 获取桌面路径
        desktop_path = os.path.expanduser("~/Desktop")

        # 创建文件的完整路径
        file_path = os.path.join(desktop_path, file_name)

        # 将DataFrame写入Excel文件
        df.to_excel(file_path, index=False)

        # 使用openpyxl加载生成的Excel文件
        wb = load_workbook(file_path)
        ws = wb.active

        # 固定第一行
        ws.freeze_panes = ws['A2']

        # 设置表头颜色为浅绿色
        header_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # 保存修改后的文件
        wb.save(file_path)

        # 弹出消息框提示用户文件已保存
        QMessageBox.about(self, "提示", f"文件已保存到桌面：{file_name}")

    #  清除tableWidget_3
    @pyqtSlot()
    def clear_tableWidget_3(self):
        self.tableWidget_3.setRowCount(0)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = VehicleTest()

    window.show()
    sys.exit(app.exec_())
